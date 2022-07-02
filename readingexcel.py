import pandas as pd
import  openpyxl

#pd.__version__




#pd.read_excel('1test1.xlsx')

#pd.read_excel('1test1.xlsx')#, index_col=0

#print(pd.read_excel('1test1.xlsx'))
"""
wb = openpyxl.load_workbook('1test1.xlsx')
print(type(wb))
sheets = wb.sheetnames
print(sheets)
print(wb.active.title)
sh1 = wb['Tabelle1']
data2=wb['Tabelle1']['A2'].value

print(data2)

#print(type(sh1))
print(sh1)
xd=sh1.cell(2,2).value
print(xd)

"""
"""
wb= openpyxl.load_workbook('1test1.xlsx')
#print(wb.active.title)
sh1 = wb[wb.active.title]
row = sh1.max_row
column = sh1.max_column
#print(row)
#print(column)

for i in range(2,row+1):
    for j in range (1,column +1 ):
        print(sh1.cell(i,j).value)
"""
class readplusexcel():

    #def __init__(self, file):
     #   self.file = file

    def dateiausgabe(file):

        wb = openpyxl.load_workbook(file)
        #wb.active würde die aktive file nutzen, könnte genutzt werden
        sh1 = wb[wb.active.title]
        row = sh1.max_row
        column = sh1.max_column
        print(row)
        print(column)

        for i in range(2, row + 1):
            for j in range(1, column + 1):
                return (sh1.cell(i, j).value)


    def dateiausgabe_column1(file):
        wb = openpyxl.load_workbook(file)
        # wb.active würde die aktive file nutzen, könnte genutzt werden
        sh1 = wb[wb.active.title]
        row = sh1.max_row
        column = sh1.max_column
        #print(row)
        #print(column)

        for i in range(2, row + 1):
            for j in range(1, 2):
                return (sh1.cell(i, j).value)



    def dateiausgabe_column2(file):
        wb = openpyxl.load_workbook(file)
        # wb.active würde die aktive file nutzen, könnte genutzt werden
        sh1 = wb[wb.active.title]
        row = sh1.max_row
        column = sh1.max_column
        #print(row)
        #print(column)

        for i in range(2, row + 1):
            for j in range(2, 3):
                return (sh1.cell(i, j).value) #vllt print


#readplusexcel.dateiausgabe('1test1.xlsx')

readplusexcel.dateiausgabe_column1('1test1.xlsx')
readplusexcel.dateiausgabe_column2('1test1.xlsx')
"""
cc= openpyxl.load_workbook('1test1.xlsx')
s2= cc[cc.active.title]                           # das ist ein konzept test bzw idee für schreiben 
row =s2.max_row
column = s2.max_column


print(s2.cell(row, column).value)

"""



